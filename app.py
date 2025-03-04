from flask import Flask, request, render_template, send_file
import pandas as pd
import os
import openpyxl
import logging
import numpy as np
from io import BytesIO
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime, timedelta

#downlaod history

def cm_to_EMU(width, height):
    return (int(width * 360000), int(height * 360000))

def write_to_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if cell.coordinate in ws.merged_cells:
        merged_range = [merged for merged in ws.merged_cells.ranges if cell.coordinate in merged][0]
        cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    cell.value = value
    return cell

def process_download_history(file_terminal_download_sharing, file_terminal_download_fms, version_a920pro, version_x990):
    def read_excel_file(file_path):
        try:
            df = pd.read_excel(file_path, header=4)
            logging.info(f"Columns in file {file_path}: {df.columns.tolist()}")
            return df
        except Exception as e:
            logging.error(f"Error membaca file {file_path}: {e}")
            return pd.DataFrame()

    # Membaca dan menggabungkan semua file
    dfs = []
    for file in file_terminal_download_sharing + file_terminal_download_fms:
        df = read_excel_file(file)
        if not df.empty:
            dfs.append(df)

    if not dfs:
        logging.error("Tidak ada data yang berhasil dibaca dari file")
        return pd.DataFrame()

    df_combined = pd.concat(dfs, ignore_index=True)

    # Filter data sesuai dengan "Download Completed start Scheduller"
    df_filtered = df_combined[
        (
            (
                (df_combined['Terminal Model'] == 'A920Pro') &
                (df_combined['App Name'] == 'A920PRO_BRIREGULAR') &
                (df_combined['Version'] == version_a920pro)
            ) |
            (
                (df_combined['Terminal Model'] == 'X990') &
                (df_combined['App Name'] == 'X990_BRIREGULAR') &
                (df_combined['Version'] == version_x990)
            )
        ) &
        (df_combined['Status'] == 'Completed')
    ]

    # Menghilangkan duplikasi berdasarkan Serial Number
    df_filtered = df_filtered.drop_duplicates(subset=['Serial Number'])

    # Mengkonversi 'Date Time' ke datetime dan mengambil tanggalnya saja
    df_filtered['Date'] = pd.to_datetime(df_filtered['Date Time']).dt.date

    # Menghitung jumlah download per hari
    daily_counts = df_filtered.groupby('Date').size().reset_index(name='Jumlah Download Completed')
    daily_counts = daily_counts.sort_values('Date')

    # Mengisi tanggal yang kosong dengan 0 download
    date_range = pd.date_range(start=daily_counts['Date'].min(), end=daily_counts['Date'].max())
    daily_counts = daily_counts.set_index('Date').reindex(date_range, fill_value=0).reset_index()
    daily_counts = daily_counts.rename(columns={'index': 'Date'})

    # Menghitung total kumulatif
    daily_counts['Total Kumulatif'] = daily_counts['Jumlah Download Completed'].cumsum()

    # Menghitung Pertumbuhan per Hari
    daily_counts['Pertumbuhan per Hari'] = daily_counts['Total Kumulatif'] - daily_counts['Jumlah Download Completed']

    # Mengubah nama kolom 'Date' menjadi 'Tanggal Download'
    daily_counts = daily_counts.rename(columns={'Date': 'Tanggal Download'})

    return daily_counts

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def add_download_history_to_worksheet(ws, download_history):
    start_row = 24  # Mulai dari baris ke-24
    start_col = 1

    # Definisikan warna merah maron
    maroon_color = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    # Tambahkan header
    headers = ["Tanggal Download", "Jumlah Download Completed", "Pertumbuhan per Hari", "Total Kumulatif"]
    for col, header in enumerate(headers, start=start_col):
        cell = write_to_cell(ws, start_row, col, header)
        cell.fill = maroon_color
        cell.font = white_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Tambahkan border ke header
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(start_col, start_col + len(headers)):
        ws.cell(row=start_row, column=col).border = border

    # Tambahkan data
    for _, row in download_history.iterrows():
        start_row += 1
        ws.cell(row=start_row, column=start_col, value=row['Tanggal Download']).number_format = 'yyyy-mm-dd'
        
        # Format Jumlah Download Completed dengan '+' di depan dan rata kanan
        jumlah_download_cell = ws.cell(row=start_row, column=start_col+1, value=row['Jumlah Download Completed'])
        jumlah_download_cell.number_format = '+#,##0;-#,##0'
        jumlah_download_cell.alignment = Alignment(horizontal='right')
        
        # Format Pertumbuhan per Hari tanpa '+' dan rata kanan
        pertumbuhan_cell = ws.cell(row=start_row, column=start_col+2, value=row['Pertumbuhan per Hari'])
        pertumbuhan_cell.number_format = '#,##0;-#,##0'
        pertumbuhan_cell.alignment = Alignment(horizontal='right')
        
        ws.cell(row=start_row, column=start_col+3, value=row['Total Kumulatif']).alignment = Alignment(horizontal='right')

        # Tambahkan border ke data
        for col in range(start_col, start_col + len(headers)):
            ws.cell(row=start_row, column=col).border = border

    # Sesuaikan lebar kolom
    for col in range(start_col, start_col + len(headers)):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in ws[column_letter][start_row-1:]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) 
        ws.column_dimensions[column_letter].width = adjusted_width

    return ws, start_row, start_row


#akhir Dwonlaod history

#Chart download history

from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
import logging
from datetime import datetime
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)
from openpyxl.drawing.spreadsheet_drawing import (
    OneCellAnchor,
    AnchorMarker
)

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU

# def add_download_history_chart(ws, start_row, end_row):
#     chart = ScatterChart()
#     chart.title = "Download History"
#     chart.style = 2
#     chart.y_axis.title = 'Jumlah EDC'
#     chart.x_axis.title = 'Tanggal'

#     # Data untuk jumlah download
#     y_values = Reference(ws, min_col=2, min_row=start_row, max_row=end_row)
#     x_values = Reference(ws, min_col=1, min_row=start_row, max_row=end_row)
#     series = Series(y_values, x_values, title="Jumlah Download Completed")
#     chart.series.append(series)

#     # Mengatur sumbu y (Jumlah EDC)
#     chart.y_axis.scaling.min = 0
#     chart.y_axis.scaling.max = 2000  # Ubah nilai maksimum menjadi 2000
#     chart.y_axis.majorUnit = 400  # Ubah unit utama menjadi 400 untuk pembagian yang lebih baik

#     # Mengatur sumbu x (Tanggal)
#     chart.x_axis.number_format = 'dd'
#     chart.x_axis.majorUnit = 1
#     chart.x_axis.tickLblPos = "low"

#     # Mengatur ukuran grafik
#     width_cm = 30
#     height_cm = 15
#     size = XDRPositiveSize2D(cm_to_EMU(width_cm), cm_to_EMU(height_cm))

#     # Menambahkan grafik ke worksheet
#     anchor = OneCellAnchor(
#         _from=AnchorMarker(col=0, colOff=0, row=end_row + 2, rowOff=0),
#         ext=size
#     )
#     ws.add_chart(chart, anchor)
    
#     return ws



# Tambahkan fungsi ini untuk menyesuaikan lebar kolom
def adjust_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width



#end chart

#Chart Buat 

import numpy as np

def create_charts(data, total_populasi, version_a920pro):
    charts = []

    # Helper function untuk menghitung nilai yang tidak aktif/belum didownload/diapply
    def calculate_inactive(active):
        return max(0, total_populasi - active)

    # 1. Data Download Tams Sharing 10.2.2.5:7000
    download_sharing = data['sharing'].get('Download Completed start Scheduller', {}).get('Total', 0)
    charts.append(create_pie_chart(
        [download_sharing, calculate_inactive(download_sharing)],
        ['Downloaded', 'Not Downloaded'],
        ['#005FAC', '#b81414'],
        'Data Download Tams Sharing\n10.2.2.5:7000\n\n'
    ))

    # 2. Data Apply Tams Sharing 10.2.2.5:7000
    apply_sharing = data['sharing'].get('Apply config', {}).get('Total', 0)
    charts.append(create_pie_chart(
        [apply_sharing, calculate_inactive(apply_sharing)],
        ['Applied', 'Not Applied'],
        ['#005FAC', '#b81414'],
        'Data Apply Tams Sharing\n10.2.2.5:7000\n\n'
    ))

    # 3. Data Download Tams FMS 10.2.30.2:7000
    download_fms = data['fms'].get('Download Completed start Scheduller', {}).get('Total', 0)
    charts.append(create_pie_chart(
        [download_fms, calculate_inactive(download_fms)],
        ['Downloaded', 'Not Downloaded'],
        ['#005FAC', '#b81414'],
        'Data Download Tams FMS\n10.2.30.2:7000\n\n'
    ))

    # 4. Data Apply Tams FMS 10.2.30.2:7000
    apply_fms = data['fms'].get('Apply config', {}).get('Total', 0)
    charts.append(create_pie_chart(
        [apply_fms, calculate_inactive(apply_fms)],
        ['Applied', 'Not Applied'],
        ['#005FAC', '#b81414'],
        'Data Apply Tams FMS\n10.2.30.2:7000\n\n'
    ))

    # 5 & 6. Data Download/Apply Tams FMS dan Sharing
    download_combined = download_sharing + download_fms
    apply_combined = apply_sharing + apply_fms
    
    charts.append(create_pie_chart(
        [download_combined, calculate_inactive(download_combined)],
        ['Downloaded', 'Not Downloaded'],
        ['#005FAC', '#b81414'],
        'Data Download Tams FMS dan Sharing\n\n'
    ))
    
    charts.append(create_pie_chart(
        [apply_combined, calculate_inactive(apply_combined)],
        ['Applied', 'Not Applied'],
        ['#005FAC', '#b81414'],
        'Data Apply Tams FMS dan Sharing\n\n'
    ))


       # 7. Data Update Aplikasi Versi
    active_transaction = data['sharing'].get('Active transaksi', {}).get('Total', 0) + data['fms'].get('Active transaksi', {}).get('Total', 0)
    downloaded = download_combined
    active_not_downloaded = max(0, active_transaction - downloaded)
    inactive = max(0, total_populasi - downloaded - active_not_downloaded)

    # Mengambil 2 digit awal dari versi aplikasi
    version_2digit = '.'.join(version_a920pro.split('.')[:2])

    charts.append(create_pie_chart(
        [downloaded, active_not_downloaded, inactive],
        ['EDC Sudah OTA', 'EDC Belum OTA', 'EDC Tidak Aktif'],
        ['#ffa500', '#005FAC', '#b81414'],
        f'Data Update Aplikasi Versi {version_2digit} Primavista \n\n '
    ))

    return charts

def create_pie_chart(sizes, labels, colors, title):
    fig, ax = plt.subplots(figsize=(4, 3))  # Mengubah ukuran menjadi 4x3 inch
    
    total = sum(sizes)
    
    def autopct_format(values):
        def my_format(pct):
            total = sum(values)
            val = int(round(pct*total/100.0))
            return f'{val:,}\n({pct:.1f}%)'
        return my_format
    
    wedges, texts, autotexts = ax.pie(sizes, 
                                      labels=labels, 
                                      colors=colors, 
                                      autopct=autopct_format(sizes),
                                      startangle=90,
                                      pctdistance=0.75)
    
    ax.axis('equal')
    
    plt.title(f"{title}", fontsize=8, fontweight='bold')  # Mengurangi ukuran font judul
    
    # Menambahkan legenda dengan nilai, menggunakan ukuran font yang lebih kecil
    legend_labels = [f'{label}: {size:,}' for label, size in zip(labels, sizes)]
    ax.legend(wedges, legend_labels, title="Kategori", loc="center left", 
              bbox_to_anchor=(1, 0, 0.5, 1), fontsize=6, title_fontsize=7)
    
    plt.setp(autotexts, size=6, weight="bold")  # Mengurangi ukuran font teks otomatis
    plt.setp(texts, size=7)  # Mengurangi ukuran font label
    
    plt.tight_layout()
    
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=300)
    img_buffer.seek(0)
    img = Image(img_buffer)
    plt.close(fig)
     # Atur ukuran gambar dalam piksel
    img.width = 380  # 4 inci * 95 DPI
    img.height = 285  # 3 inci * 95 DPI
    
    plt.close(fig)
    return img

from openpyxl.styles import Font

def add_specific_charts_to_worksheet(ws, charts):
    chart_positions = [
        ('H', 2),  # 1. Data Download Tams Sharing 10.2.2.5:7000
        ('O', 2),  # 2. Data Apply Tams Sharing 10.2.2.5:7000
        ('H', 19), # 3. Data Download Tams FMS 10.2.30.2:7000
        ('O', 19), # 4. Data Apply Tams FMS 10.2.30.2:7000
        ('H', 35), # 5. Data Download Tams FMS dan Sharing
        ('O', 35), # 6. Data Apply Tams FMS dan Sharing
        ('H', 51), # 7. Data Update Aplikasi Versi
    ]

    titles = [
        "1. Data Download Tams Sharing 10.2.2.5:7000",
        "2. Data Apply Tams Sharing 10.2.2.5:7000",
        "3. Data Download Tams FMS 10.2.30.2:7000",
        "4. Data Apply Tams FMS 10.2.30.2:7000",
        "5. Data Download Tams FMS dan Sharing",
        "6. Data Apply Tams FMS dan Sharing",
        "7. Data Update Aplikasi Versi"
    ]

    for i, (chart, (col, row)) in enumerate(zip(charts, chart_positions)):
        # Menambahkan judul untuk setiap chart
        ws.cell(row=row-1, column=ord(col)-64).value = titles[i]
        ws.cell(row=row-1, column=ord(col)-64).font = Font(bold=True)

        # Menambahkan gambar
        ws.add_image(chart, f'{col}{row}')
        
        # Menyesuaikan ukuran sel
        img_width_cm = chart.width / 37.795275591  # Konversi piksel ke cm
        img_height_cm = chart.height / 37.795275591
        
        ws.column_dimensions[col].width = img_width_cm / 0.748031496  # Konversi cm ke lebar kolom Excel
        
        # Menyesuaikan tinggi baris untuk chart
        for j in range(15):  # Mengasumsikan setiap chart membutuhkan 15 baris
            ws.row_dimensions[row + j].height = (img_height_cm / 15) / 0.035  # Konversi cm ke tinggi baris Excel

    return ws
#Akhir Chart Buat


def add_dataframe_to_worksheet(ws, df, start_row, start_col):
    # Add headers
    for col, header in enumerate(df.columns, start=start_col):
        write_to_cell(ws, start_row, col, header)
    
    # Add data
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row+1):
        for c_idx, value in enumerate(row, start=start_col):
            write_to_cell(ws, r_idx, c_idx, value)

app = Flask(__name__)
logging.basicConfig(level=logging.INFO)

def read_excel_file(file_path, skiprows):
    try:
        return pd.read_excel(file_path, skiprows=skiprows)
    except Exception as e:
        logging.error(f"Error reading file {file_path}: {e}")
        return None

def process_a920_data(df, version):
    filtered_a920pro = df[
        (df['Terminal Model'] == 'A920Pro') &
        (df['App Name'] == 'A920PRO_BRIREGULAR') &
        (df['Version'] == version) &
        (df['Status'] == 'Completed')
    ]
    return filtered_a920pro['Serial Number'].nunique()

def process_x990_data(df, version):
    filtered_x990 = df[
        (df['Terminal Model'] == 'X990') &
        (df['App Name'] == 'X990_BRIREGULAR') &
        (df['Version'] == version) &
        (df['Status'] == 'Completed')
    ]
    return filtered_x990['Serial Number'].nunique()


def process_files(file_paths, version_a920pro, version_x990):
    df_a920 = read_excel_file(file_paths['file_terminal_version_a920'][0], 8)
    df_x990 = read_excel_file(file_paths['file_terminal_version_x990'][0], 8)
    
    # Membaca dan menggabungkan semua file terminal download
    df_terminal_download_list = [read_excel_file(file, 4) for file in file_paths['file_terminal_download']]
    df_terminal_download = pd.concat(df_terminal_download_list, ignore_index=True)
    
    df_data_aktif = read_excel_file(file_paths['file_data_aktif'][0], 0)

    if any(df is None for df in [df_a920, df_x990] + df_terminal_download_list + [df_data_aktif]):
        return None, "Error: One or more required files are missing or invalid"


    a920_populasi = df_a920['Serial Number'].nunique()
    x990_populasi = df_x990['Serial Number'].nunique()

    # Perhitungan Download Completed start Scheduller
    a920_download_completed = df_terminal_download[
        (df_terminal_download['Terminal Model'] == 'A920Pro') &
        (df_terminal_download['App Name'] == 'A920PRO_BRIREGULAR') &
        (df_terminal_download['Version'] == version_a920pro) &
        (df_terminal_download['Status'] == 'Completed')
    ]['Serial Number'].nunique()

    x990_download_completed = df_terminal_download[
        (df_terminal_download['Terminal Model'] == 'X990') &
        (df_terminal_download['App Name'] == 'X990_BRIREGULAR') &
        (df_terminal_download['Version'] == version_x990) &
        (df_terminal_download['Status'] == 'Completed')
    ]['Serial Number'].nunique()

    # Perhitungan Apply config
    apply_config_a920pro = df_a920[
        (df_a920['APP Name'] == 'A920PRO_BRIREGULAR') &
        (
            (df_a920['Actual APP Version'] == version_a920pro) |
            (df_a920['Actual APP Version'] == '0.0.0.0') |
            (df_a920['Actual APP Version'].isnull())
        )
    ]['Serial Number'].nunique()

    apply_config_x990 = df_x990[
        (df_x990['APP Name'] == 'X990_BRIREGULAR') &
        (
            (df_x990['Actual APP Version'] == version_x990) |
            (df_x990['Actual APP Version'] == '0.0.0.0') |
            (df_x990['Actual APP Version'].isnull())
        )
    ]['Serial Number'].nunique()

    df_data_aktif['FSN'] = df_data_aktif['FSN'].fillna('').astype(str)
    a920_active_transaksi = df_data_aktif[df_data_aktif['FSN'].str.startswith('185')]['FSN'].nunique()
    x990_active_transaksi = df_data_aktif[df_data_aktif['FSN'].str.startswith('V1E')]['FSN'].nunique()

    total_populasi_tams = a920_populasi + x990_populasi
    total_download_completed = a920_download_completed + x990_download_completed
    total_apply_config = apply_config_a920pro + apply_config_x990
    total_active_transaksi = a920_active_transaksi + x990_active_transaksi

    result_df = pd.DataFrame({
        'Type': ['A920pro', 'X990', 'Total'],
        'Populasi Tams': [a920_populasi, x990_populasi, total_populasi_tams],
        'Download Completed start Scheduller': [a920_download_completed, x990_download_completed, total_download_completed],
        'Apply config': [apply_config_a920pro, apply_config_x990, total_apply_config],
        'Active transaksi': [a920_active_transaksi, x990_active_transaksi, total_active_transaksi]
    })

    return result_df

expected_files = {
    'sharing': ['file_data_aktif', 'file_terminal_download', 'file_terminal_version_a920', 'file_terminal_version_x990'],
    'fms': ['file_data_aktif', 'file_terminal_download', 'file_terminal_version_a920', 'file_terminal_version_x990']
}

logging.basicConfig(level=logging.DEBUG)

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        logging.info("Received POST request")
        logging.debug(f"Form data: {request.form}")
        logging.debug(f"Files received: {request.files}")
        
        try:
            file_paths = {'sharing': {}, 'fms': {}}
            
            for data_type, file_list in expected_files.items():
                logging.info(f"Processing {data_type} files")
                for file_key in file_list:
                    files = request.files.getlist(f"{file_key}_{data_type}")
                    logging.info(f"Files for {file_key}_{data_type}: {[f.filename for f in files]}")
                    
                    if files:
                        for file in files:
                            file_path = os.path.join('uploads', file.filename)
                            file.save(file_path)
                            logging.info(f"Saved file: {file_path}")
                            if file_key not in file_paths[data_type]:
                                file_paths[data_type][file_key] = []
                            file_paths[data_type][file_key].append(file_path)
                    else:
                        logging.error(f"Missing required file: {file_key} for {data_type}")
                        return f"Error: Missing required file: {file_key} for {data_type}", 400

            logging.info("All required files received successfully")

            versions = {
                'sharing': {
                    'a920pro': request.form.get('version_a920pro_sharing'),
                    'x990': request.form.get('version_x990_sharing')
                },
                'fms': {
                    'a920pro': request.form.get('version_a920pro_fms'),
                    'x990': request.form.get('version_x990_fms')
                }
            }
            logging.info(f"Versions: {versions}")

            # Get total_populasi from form
            total_populasi = int(request.form.get('total_populasi', 0))
            output_filename = request.form.get('output_filename', 'Default')

            logging.info("Starting file processing")

            wb = openpyxl.Workbook()
            ws = wb.active

            results = {}
            for data_type, title in [('sharing', 'TAMS SHARING 10.2.2.5:7000'), 
                                    ('fms', 'TAMS FMS 10.2.30.2:7000')]:
                result = process_files(
                    file_paths[data_type], 
                    versions[data_type]['a920pro'], 
                    versions[data_type]['x990']
                )

                if result is None:
                    return f"Error processing data for {data_type}", 400
                
                results[data_type] = result

                if ws.max_row > 1:
                    ws.append([])

                ws.append([title])
                title_cell = ws.cell(row=ws.max_row, column=1)
                title_cell.font = Font(size=16, bold=True)
                title_cell.alignment = Alignment(horizontal='center')
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)

                add_dataframe_to_worksheet(ws, results[data_type], ws.max_row + 1, 1)
                last_row = ws.max_row
            
            start_row = 1

            for data_type, title in [('sharing', 'TAMS SHARING 10.2.2.5:7000'),
                                    ('fms', 'TAMS FMS 10.2.30.2:7000')]:
                result = process_files(
                    file_paths[data_type],
                    versions[data_type]['a920pro'],
                    versions[data_type]['x990']
                )

                if result is None:
                    return f"Error processing data for {data_type}", 400

                results[data_type] = result

                def write_to_cell(ws, row, col, value):
                    cell = ws.cell(row=row, column=col)
                    if cell.coordinate in ws.merged_cells:
                        merged_range = [merged for merged in ws.merged_cells.ranges if cell.coordinate in merged][0]
                        cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    cell.value = value
                    return cell

                title_cell = write_to_cell(ws, start_row, 1, title)
                title_cell.font = Font(size=16, bold=True)
                title_cell.alignment = Alignment(horizontal='center')
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)

                add_dataframe_to_worksheet(ws, results[data_type], start_row + 1, 1)
                start_row += len(results[data_type]) + 3  # +3 for title, header, and space

            # Add "TAMS FMS dan SHARING" section
            ws.cell(row=13, column=1, value="TAMS FMS dan SHARING")
            title_cell = ws.cell(row=13, column=1)
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=13, start_column=1, end_row=13, end_column=5)

            # Combine data from TAMS Sharing and TAMS FMS
            combined_data = results['sharing'].set_index('Type') + results['fms'].set_index('Type')
            combined_data = combined_data.reset_index()

            add_dataframe_to_worksheet(ws, combined_data, 14, 1)

            # Process download history
            file_terminal_download_sharing = file_paths['sharing']['file_terminal_download']
            file_terminal_download_fms = file_paths['fms']['file_terminal_download']

            if file_terminal_download_sharing and file_terminal_download_fms:
                try:
                    download_history = process_download_history(
                        file_terminal_download_sharing,
                        file_terminal_download_fms,
                        versions['sharing']['a920pro'],
                        versions['sharing']['x990']
                    )
                    if not download_history.empty:
                        ws, start_row, end_row = add_download_history_to_worksheet(ws, download_history)
                        # ws = add_download_history_chart(ws, start_row, end_row)

                        # Menyesuaikan lebar kolom
                        adjust_column_width(ws)
                    else:
                        logging.warning("Download history is empty")
                        write_to_cell(ws, last_row+1, 1, "Tidak ada data download history yang memenuhi kriteria")
                        last_row += 1
                except Exception as e:
                    logging.error(f"Error processing download history: {e}")
                    ws.cell(row=last_row+1, column=1, value="Error dalam memproses download history")
                    last_row += 1
            else:
                logging.error("One or both download history files are missing")
                ws.cell(row=last_row+1, column=1, value="File download history tidak lengkap")
                last_row += 1

            # Adjust column widths
            adjust_column_width(ws)

            # Add total_populasi
            last_row += 1
            write_to_cell(ws, last_row, 1, "Total Populasi")
            write_to_cell(ws, last_row, 2, total_populasi)

            # Membuat data untuk charts
            data = {
                'sharing': results['sharing'].set_index('Type').to_dict(),
                'fms': results['fms'].set_index('Type').to_dict(),
            }

            # Membuat charts
            charts = create_charts(data, total_populasi, versions['sharing']['a920pro'])

            # Menambahkan charts ke worksheet
            add_specific_charts_to_worksheet(ws, charts)

            # Simpan workbook ke dalam BytesIO object
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Delete uploaded files
            for data_type in file_paths:
                for file_key in file_paths[data_type]:
                    for file_path in file_paths[data_type][file_key]:
                        try:
                            os.remove(file_path)
                            logging.info(f"Deleted file: {file_path}")
                        except Exception as e:
                            logging.error(f"Error deleting file: {e}")

            # Kirim file sebagai attachment
            return send_file(
                output,
                as_attachment=True,
                download_name=f"File Data OTA BRI FMS {output_filename}.xlsx",
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            logging.exception(f"Unexpected error occurred: {e}")
            # Attempt to delete any remaining uploaded files
            for data_type in file_paths:
                for file_key in file_paths[data_type]:
                    for file_path in file_paths[data_type][file_key]:
                        try:
                            if os.path.exists(file_path):
                                os.remove(file_path)
                                logging.info(f"Deleted file during error handling: {file_path}")
                        except Exception as delete_error:
                            logging.error(f"Error deleting file during error handling: {delete_error}")
            return f"An unexpected error occurred: {str(e)}", 500

    return render_template('index.html')

if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(host='0.0.0.0', port=5112, debug=True)
                    
        